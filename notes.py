from openpyxl import load_workbook

# Load  workbook and specify sheet from Excel file containing notes
notes_info = load_workbook(filename='CLIENT INFORMATION.xlsx')
sheet_name = notes_info.sheetnames[2]
sheet = notes_info[sheet_name]


def notes():
    notes_dict = {}
    # Iterate through each entry/cell to search notes and store client name and related notes in dictionary as k, v
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and not str(cell.value).startswith("Client") and not str(cell.value).startswith(","):
                if cell.comment:  # Check if the cell has a note
                    note_text = cell.comment.text.strip()
                    notes_dict[cell.value.strip()] = note_text
                else:
                    notes_dict[cell.value.strip()] = "No Notes."

    return notes_dict
